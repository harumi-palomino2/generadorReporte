import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from funciones import quitar_ceros_en_exp, eliminar_fe_en_situacion, descomponer_columna_exp, generar_ID

st.set_page_config(page_title="Procesador de Excel", layout="centered")
st.title("📊 Procesador de archivos Excel")

st.markdown("Sube un archivo `.xlsx`, elige una hoja y aplica las funciones deseadas.")

# --- SUBIR ARCHIVO ---
archivo_excel = st.file_uploader("📤 Subir archivo Excel", type=["xlsx"])

if archivo_excel:
    nombre_archivo_original = archivo_excel.name
    wb = load_workbook(filename=archivo_excel)
    hojas = wb.sheetnames
    hoja = st.selectbox("📑 Selecciona una hoja", hojas)

    if hoja:
        ws_origen = wb[hoja]

        # Si ya existe la hoja 'procesado', eliminarla
        if "procesado" in wb.sheetnames:
            del wb["procesado"]

        # Copiar hoja original y renombrar como 'procesado'
        ws = wb.copy_worksheet(ws_origen)
        ws.title = "procesado"

        # --- FUNCIONES A APLICAR ---
        st.markdown("### 🔧 Selecciona funciones a aplicar:")
        f_ceros = st.checkbox("Quitar ceros en Exp.")
        f_situacion = st.checkbox("Unificar situación (fe:)")
        f_descomponer = st.checkbox("Descomponer columna Exp.")
        f_id = st.checkbox("Generar ID")

        if st.button("✅ Procesar"):
            try:
                funciones_aplicadas = []

                with st.spinner("Procesando... ⏳"):
                    if f_ceros:
                        quitar_ceros_en_exp(ws)
                        funciones_aplicadas.append("Quitar ceros en Exp.")
                    if f_situacion:
                        eliminar_fe_en_situacion(ws)
                        funciones_aplicadas.append("Unificar situación (fe:)")
                    if f_descomponer:
                        descomponer_columna_exp(ws)
                        funciones_aplicadas.append("Descomponer columna Exp.")
                    if f_id:
                        generar_ID(ws)
                        funciones_aplicadas.append("Generar ID")

                    if not funciones_aplicadas:
                        st.warning("⚠️ Debes seleccionar al menos una función.")
                    else:
                        # Guardar archivo modificado en memoria
                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)

                        st.success("✅ Procesamiento completado.")
                        st.markdown("Funciones aplicadas:")
                        st.markdown("• " + "\n• ".join(funciones_aplicadas))

                        st.download_button(
                            label="📥 Descargar archivo con hoja procesada",
                            data=output,
                            file_name=nombre_archivo_original,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            except Exception as e:
                st.error(f"❌ Error: {e}")
