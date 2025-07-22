import pandas as pd
import re
from openpyxl import load_workbook


def quitar_ceros_en_exp(ws):
    headers = [c.value for c in ws[3]]
    if "Exp." not in headers:
        raise ValueError("No se encontró la columna 'Exp.' en la hoja seleccionada.")

    col_idx = headers.index("Exp.") + 1

    for fila in range(4, ws.max_row + 1):
        celda = ws.cell(row=fila, column=col_idx)
        if celda.value:
            celda.value = str(celda.value).lstrip("0")


def eliminar_fe_en_situacion(ws):
    headers = [c.value for c in ws[3]]
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]

    if "situación del expediente" in headers_lower:
        col_idx = headers_lower.index("situación del expediente") + 1

        def limpiar_texto(texto):
            if texto is None:
                return texto
            return re.sub(r'\s*fe:.*?\)', ')', str(texto), flags=re.IGNORECASE)

        for fila in range(4, ws.max_row + 1):
            celda = ws.cell(row=fila, column=col_idx)
            celda.value = limpiar_texto(celda.value)
    else:
        raise ValueError("No se encontró la columna 'Situación del Expediente' en la hoja.")



def descomponer_columna_exp(ws):
    headers = [c.value for c in ws[3]]
    exp_idx = headers.index("Exp.")
    col_idx = exp_idx + 1

    nuevos_datos = []

    for fila in range(4, ws.max_row + 1):
        celda = ws.cell(row=fila, column=col_idx)
        s = str(celda.value).strip() if celda.value else ""
        m = re.match(r"^0*(\d+)([A-Z]{0,2})?$", s)
        if m:
            num, letters = m.group(1), m.group(2) or ""
            numero = int(num)
            exp_str = num
            if len(num) > 3:
                p1, p2 = int(num[:-3]), int(num[-3:])
            else:
                p1, p2 = "", int(num)
            disg = letters or "1"
        else:
            exp_str, p1, p2, disg = "", "", "", ""
        nuevos_datos.append([exp_str, p1, p2, disg])

    pos = col_idx + 1
    ws.insert_cols(pos, amount=4)
    for i, h in enumerate(["exp.", "1", "2", "disg"]):
        ws.cell(row=3, column=pos + i, value=h)

    for row, datos in enumerate(nuevos_datos, start=4):
        for i, val in enumerate(datos):
            col = pos + i
            ws.cell(row=row, column=col, value=val)


def generar_ID(ws):
    # Leer encabezados originales
    headers = [c.value for c in ws[3]]
    if "Exp." not in headers or "Año Exp." not in headers:
        raise ValueError("No se encontró 'Exp.' o 'Año Exp.' en la hoja.")

    col_exp = headers.index("Exp.") + 1
    col_id = col_exp + 1

    # Insertar columna para "ID"
    ws.insert_cols(col_id)
    ws.cell(row=3, column=col_id, value="ID")

    # ⚠️ Recalcular índice de Año Exp. porque ahora se movió a la derecha
    headers_actualizados = [c.value for c in ws[3]]
    if "Año Exp." not in headers_actualizados:
        raise ValueError("Después de insertar ID, no se encontró 'Año Exp.'.")

    col_anio = headers_actualizados.index("Año Exp.") + 1

    for fila in range(4, ws.max_row + 1):
        exp = ws.cell(row=fila, column=col_exp).value
        anio = ws.cell(row=fila, column=col_anio).value
        if exp is not None and anio is not None:
            ws.cell(row=fila, column=col_id, value=f"{exp}{anio}")
        else:
            ws.cell(row=fila, column=col_id, value="")
